import logging
import os
import threading
import time

from openpyxl import load_workbook

from app.core.config import settings

logger = logging.getLogger(__name__)

# Fields used to build the lookup key — must match column names in the mappings sheet.
# "dalecode" and "plancode" are NOT included: dalecode is derived, plancode is the result.
FIELDS = [
    "underwriting",
    "corecover",
    "psych",
    "gpreferred",
    "hospital_list",
    "opticaldental",
    "6week",
    "excess",
    "benefitreduction",
    "oplimit",
    "islands",
]

# Normalise field names coming from the values sheet (handle typos / casing differences)
FIELD_NAME_MAP = {
    "hostptal_list": "hospital_list",   # typo in spreadsheet
    "Islands":       "islands",
    "6Week":         "6week",
    "6WEEK":         "6week",
}

# ── Thread-safe state ─────────────────────────────────────────────────────────
_STATE_LOCK = threading.Lock()
_state: dict = {
    "lookup": {},
    "all_rows": [],
    "field_values": {},
    "version_info": {},
    "load_info": {},
}


def _get_state() -> dict:
    return _state


# ── Loader ────────────────────────────────────────────────────────────────────

def load() -> None:
    """
    Load (or reload) all sheets from the xlsx file into memory.
    Thread-safe: builds new state locally, then atomically swaps _state.
    """
    global _state
    path = settings.DATA_FILE

    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Plancode mappings file not found at {path}. "
            "Upload plancode_mappings.xlsx via the admin UI or place it in the data directory."
        )

    wb = load_workbook(filename=path, read_only=True, data_only=True)

    # ── Sheet: mappings ───────────────────────────────────────────────────────
    ws_map = wb["mappings"]
    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(ws_map.iter_rows(min_row=1, max_row=1))
    ]
    lookup: dict[tuple, str] = {}
    rows: list[dict] = []

    for row in ws_map.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        d = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
        key = tuple(d.get(f, "") for f in FIELDS)
        lookup[key] = d.get("plancode", "").strip()
        rows.append(d)

    # ── Sheet: values (long format: field | value | label) ────────────────────
    ws_val = wb["values"]
    field_values: dict[str, list[dict]] = {}

    for row in ws_val.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        raw_field = str(row[0]).strip()
        field_name = FIELD_NAME_MAP.get(raw_field, raw_field)

        value = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        label = str(row[2]).strip() if len(row) > 2 and row[2] is not None else value

        if not value:
            continue

        if field_name not in field_values:
            field_values[field_name] = []

        # Deduplicate by value — keeps first occurrence
        if not any(item["value"] == value for item in field_values[field_name]):
            field_values[field_name].append({"value": value, "label": label or value})

    # ── Sheet: version ────────────────────────────────────────────────────────
    ws_ver = wb["version"]
    ver_headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(ws_ver.iter_rows(min_row=1, max_row=1))
    ]
    version_info: dict = {}
    for row in ws_ver.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            version_info = {
                ver_headers[i]: (str(v).strip() if v is not None else "")
                for i, v in enumerate(row)
            }

    wb.close()

    load_info = {
        "path": path,
        "filename": os.path.basename(path),
        "row_count": len(rows),
        "loaded_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "file_size_kb": round(os.path.getsize(path) / 1024, 1),
    }

    new_state = {
        "lookup": lookup,
        "all_rows": rows,
        "field_values": field_values,
        "version_info": version_info,
        "load_info": load_info,
    }

    with _STATE_LOCK:
        _state = new_state  # single reference swap — atomic in CPython

    logger.info("Plancode mappings loaded", extra={"row_count": len(rows), "path": path})


# ── Public accessors ──────────────────────────────────────────────────────────

def resolve(inputs: dict) -> str | None:
    """Look up a plancode from a dict of field → value."""
    key = tuple(str(inputs.get(f, "")).strip() for f in FIELDS)
    return _get_state()["lookup"].get(key)


def get_all_rows() -> list[dict]:
    return _get_state()["all_rows"]


def get_load_info() -> dict:
    return _get_state()["load_info"]


def get_field_values() -> dict[str, list[dict]]:
    """Returns {field: [{value, label}, ...]} for all fields in the values sheet."""
    return _get_state()["field_values"]


def get_version_info() -> dict:
    return _get_state()["version_info"]
