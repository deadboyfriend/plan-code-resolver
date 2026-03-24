import os
import pytest
from openpyxl import Workbook
from fastapi.testclient import TestClient


@pytest.fixture(scope="session", autouse=True)
def test_data_file(tmp_path_factory):
    """Create a minimal xlsx with known data and point DATA_FILE at it."""
    tmp = tmp_path_factory.mktemp("data")
    path = tmp / "test_mappings.xlsx"

    wb = Workbook()

    # mappings sheet — one row of known test data
    ws_map = wb.active
    ws_map.title = "mappings"
    ws_map.append([
        "plancode", "dalecode",
        "underwriting", "corecover", "psych", "gpreferred",
        "hospital_list", "opticaldental", "sixweek", "excess",
        "benefitreduction", "oplimit", "islands",
    ])
    ws_map.append([
        "TST", "NMORIY1a2a3a4a5a6a7a8a9a",
        "NMORI", "Y", "1a", "2a", "3a", "4a", "5a", "6a", "7a", "8a", "9a",
    ])

    # values sheet — label/value pairs for each field
    ws_val = wb.create_sheet("values")
    ws_val.append(["field", "value", "label"])
    for row in [
        ("underwriting",    "NMORI", "No Moratorium"),
        ("corecover",       "Y",     "included"),
        ("psych",           "1a",    "28 days"),
        ("gpreferred",      "2a",    "Yes"),
        ("hospital_list",   "3a",    "Key"),
        ("opticaldental",   "4a",    "Yes"),
        ("sixweek",         "5a",    "Yes"),
        ("excess",          "6a",    "50"),
        ("benefitreduction","7a",    "Yes"),
        ("oplimit",         "8a",    "0"),
        ("islands",         "9a",    "N/A"),
    ]:
        ws_val.append(row)

    # version sheet
    ws_ver = wb.create_sheet("version")
    ws_ver.append(["version", "change_date", "effective_date", "author", "description"])
    ws_ver.append(["1.0", "2026-01-01", "2026-01-01", "test", "test data"])

    wb.save(path)
    os.environ["DATA_FILE"] = str(path)
    return path


@pytest.fixture(scope="session")
def client(test_data_file):
    from app.services import csv_loader
    csv_loader.load()
    from app.main import app
    with TestClient(app) as c:
        yield c
